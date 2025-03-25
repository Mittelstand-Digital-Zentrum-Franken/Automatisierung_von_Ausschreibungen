import pandas as pd
import math
from typing import Dict, List, Optional, Union, Tuple
from openpyxl import Workbook
from io import BytesIO
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font, NumberFormatDescriptor
from openpyxl.utils.dataframe import dataframe_to_rows
import json
import spacy
from spacy.training.example import Example
from transformers import pipeline


def read_excel_file(file_name: str,
                    sheet_name: str ="Sheet1") -> Optional[pd.DataFrame]:
    """Reads an Excel file and returns its content as a pandas DataFrame.
    
    Parameters:
        file_name : path to Excel file to be read.
        sheet_name : name of sheet within the Excel file to load. 
        pd.DataFrame
    Returns: DataFrame containing the contents of the specified sheet.
    
    Raises: Exception: Prints an error message if the file cannot be read.
    """

    try:
        df_upload = pd.read_excel(file_name, sheet_name=sheet_name)
        return df_upload
    except Exception as e:
        print("Fehler beim Einlesen der Excel-Datei:", e)
        return None


def select_data(df_upload: pd.DataFrame) -> pd.DataFrame:
    """Preprocesses a DataFrame by removing empty rows and columns and selecting specific columns.

    Parameters:
        df_upload: Data Frame with data of xls File.

    Returns:
        df_selected: A new DataFrame with the necessary columns.
    """

    # remove empty rows and columns
    df_upload = df_upload.dropna(how='all').reset_index(drop=True)
    df_upload = df_upload.dropna(axis=1, how='all')

    # select and rename colums
    df_selected = df_upload[['Unnamed: 1', 'Unnamed: 5', 'Unnamed: 20', 'Unnamed: 27']].copy()
    df_selected = df_selected.rename(columns={'Unnamed: 1': 'amount',
                                              'Unnamed: 5': 'description',
                                              'Unnamed: 20': 'single_price',
                                              'Unnamed: 27': 'total_price'})
    
    return df_selected


def replace_nan_inf(data: List[Dict[str, Union[str, float]]]) ->List[Dict[str, Union[str, float]]]:
    """Replace NaN and infinite values with 0."""
    if isinstance(data, list):
        return [replace_nan_inf(item) for item in data]
    elif isinstance(data, dict):
        return {key: replace_nan_inf(value) for key, value in data.items()}
    elif isinstance(data, (int, float)) and (math.isnan(data) or math.isinf(data)):
        return 
    return data


def extract_information_through_positionnumbers(df_cleaned: pd.DataFrame) -> List[Dict[str, Union[str, float]]]:
    """Extracts structured information from a DataFrame based on position numbers and saving each block as a dictionary.

    Parameters:
        df_cleaned: 'description', 'amount', 'single_price', and 'total_price' of furniture.

    Returns:
        blocks: A list of dictionaries, each representing an extracted data block of a furniture.
    """

    # list to save extracted blocks in dictionaries
    blocks = []

    # temporary variables
    current_pos = None
    current_amount = None
    current_single_price = None
    current_total_price = None
    current_articlenumber = None
    current_block = []

    # iterate through the data frame
    for idx, row in df_cleaned.iterrows():
        description = row['description']
        amount = row['amount']
        single_price = row['single_price']
        total_price = row['total_price']
        
        # Check if a new position ("Pos") was found
        if description and "Pos" in str(description):
            # If there is an active block, save as dictionary
            if current_block:
                blocks.append({
                    'position': current_pos,
                    'articlenumber': current_articlenumber,
                    'description': current_block,
                    'amount': current_amount,
                    'single_price': current_single_price,
                    'total_price': current_total_price
                })
            
            # set new position
            current_pos = description
            current_block = []
            current_amount = None
            current_single_price = None
            current_total_price = None

            # save next row as "articlenumber"
            if idx + 1 < len(df_cleaned):
                next_row = df_cleaned.iloc[idx + 1]
                current_articlenumber = next_row['description']
                continue

        # Add description to block if available
        if description and not "Pos" in str(description):
            # Only add the description if it is not an item number
            if current_articlenumber and description != current_articlenumber:
                current_block.append(description)

        # Update quantity, single_price and total_price for the block if available
        # if pd.notna(amount) and current_amount is None:
        if pd.notna(amount):
            current_amount = amount
        if pd.notna(single_price):
            current_single_price = single_price
        if pd.notna(total_price):
            current_total_price = total_price

    # Save last block if it exists
    if current_block:
        blocks.append({
            'position': current_pos,
            'articlenumber': current_articlenumber,
            'description': current_block,
            'amount': current_amount,
            'single_price': current_single_price,
            'total_price': current_total_price
        })
    
    # Replace nan values
    # blocks = [replace_nan_inf(block) for block in blocks]

    # Konvertieren der Formatierungen
    for block in blocks:
        # Sting zu Floats
        for key in ['single_price', 'total_price']:
            if block[key]:
                # Tausenderpunkte entfernen und Komma in Punkt umwandeln
                block[key] = float(block[key].replace('.', '').replace(',', '.'))
            else:
                block[key] = 0.0

        # Integers zu Strings
        for key in ['description']:
            block[key] = [str(item) for item in block[key]]

        # # Beschreibung zu einem String umwandeln
        # block['description'] = " ".join(block['description'])

    return blocks


def load_and_set_up():
    """Load the necessary models and files.

    Returns:
        nlp: Spracy model fpr NER
        mapping: mapping table to change NER words
    """
    # Load models
    nlp = spacy.load("de_core_news_lg")

    # Load mapping file
    with open("mapping.json", "r", encoding="utf-8") as file:
        mapping = json.load(file)

    # Create Entity Ruler Pipeline
    ruler = nlp.add_pipe("entity_ruler", before="ner")

    # Create pattern
    patterns = [{"label": "NEUTRAL", "pattern": key} for key in mapping.keys()]
    ruler.add_patterns(patterns)

    return nlp, mapping


def find_similar_term(term, known_terms, nlp):
    term_doc = nlp(term)
    best_match = None
    best_similarity = 0.0

    for known in known_terms:
        known_doc = nlp(known)
        similarity = term_doc.similarity(known_doc)
        if similarity > best_similarity:
            best_similarity = similarity
            best_match = known

    return best_match if best_similarity > 0.7 else None



def neutralize_entities_in_text(text: str,
                                mapping: Dict[str, str],
                                nlp) -> Tuple[str, str]:
    """Neutralise the entities if possible

    Args:
        text: Text to neutralise
        mapping: Dictionary with the known Words to neutralise
        nlp: Spacy model

    Returns:
        neutral_description: Neutralise text
        ersetzungen: Only the neutralised word
    """
    doc = nlp(text)
    neutral_description = text
    ersetzungen = {}

    for ent in doc.ents:
        original_word = ent.text
        replaced_word = ""

        # Prüfen, ob ein Begriff in der Mapping-Tabelle vorhanden ist
        if ent.label_ == "NEUTRAL" and original_word in mapping:
            replaced_word = mapping[original_word]
        else:
            # Ähnlichkeitsprüfung, falls kein passender Begriff in der Mapping-Tabelle vorhanden ist
            similar_term = find_similar_term(original_word, mapping.keys(), nlp)
            if similar_term:
                replaced_word = mapping[similar_term]
            else:
                # Falls auch keine Ähnlichkeit gefunden wird, als 'zu bearbeiten' markieren
                replaced_word = "zu bearbeiten"

        # Ersetze das Wort im Text
        neutral_description = neutral_description.replace(original_word, replaced_word)

        # Speichere die Ersetzung im Dictionary
        ersetzungen[original_word] = replaced_word

    return neutral_description, ersetzungen


def neutraliz_blocks(blocks: List[Dict[str, Union[str, float]]]) -> List[Dict[str, Union[str, float]]]:
    """Add the neutralised blocks.

    Args:
        blocks: A list of dictionaries, each representing an extracted data block of a furniture.

    Returns:
        blocks: A list of dictionaries, each representing an extracted data block of a furniture and the neutral text.
    """
    # Initialise Spacy
    nlp, mapping = load_and_set_up()

    for block in blocks:
        # Add new keys
        block["neutralise"] = []
        block["neutral_description"] = ""
        block["switch"] = {}

        # NER
        description = " ".join(block['description']) if isinstance(block['description'], list) else block['description']

        # Neutralise
        neutral_description, ersetzungen = neutralize_entities_in_text(description, mapping, nlp)
        # Add to dictionary
        block["neutralise"].extend(ersetzungen.keys())
        block["neutral_description"] = neutral_description
        block["switch"] = ersetzungen
        print("neutralisierung durchlaufen")

    return blocks


def create_cost_estimate_intern(list_of_extracted_information: List[Dict[str, Union[str, float]]]) -> pd.DataFrame:
    """Create a intern cost estimate DataFrame from extracted information.

    Parameters:
        list_of_extracted_information: necessary information of furniture.

    Returns:
        df_cost_estimate_intern: finish Data Frame for the intern cost estimate.
    """

    # create a empty list
    new_rows_intern = []

    # iterating over the data and adding the values ​​to the list
    for entry in list_of_extracted_information:
        if entry['position'] is not None:
            # prepare values
            pos = entry['position'].replace('Pos. ', '')
            articlenumber = entry['articlenumber']
            amount = entry['amount']
            single_price = entry['single_price']
            total_price = entry['total_price']

            # convert each element in description to string
            if isinstance(entry['description'], list):
                description = "\n".join(str(desc) for desc in entry['description']) + "\n"
            else:
                description = str(entry['description']) + "\n"

            # add new row to the list
            new_rows_intern.append({
                'Pos:': pos,
                'Hersteller': None,
                'Bezeichnung': description,
                'Artikelnummer': articlenumber,
                'Menge': amount,
                'Einzelpreis / EUR netto': single_price,
                'Gesamtpreis / EUR netto': total_price
            })

    # convert into an data frame
    df_cost_estimate_intern = pd.DataFrame(new_rows_intern)

    return df_cost_estimate_intern


def create_cost_estimate_extern(list_of_extracted_information: List[Dict[str, Union[str, float]]]) -> pd.DataFrame:
    """Create a intern cost estimate DataFrame from extracted information.

    Parameters:
        list_of_extracted_information: necessary information of furniture.

    Returns:
        df_cost_estimate_intern: finish Data Frame for the intern cost estimate.
    """

    # create a empty list
    new_rows_extern = []

    # iterating over the data and adding the values ​​to the list
    for entry in list_of_extracted_information:
        if entry['position'] is not None:
            # prepare values
            pos = entry['position'].replace('Pos. ', '')
            amount = entry['amount']

            # convert each element in description to string
            if isinstance(entry['neutral_description'], list):
                description = "\n".join(str(desc) for desc in entry['neutral_description']) + "\n"
            else:
                description = str(entry['neutral_description']) + "\n"

            # add new row to the list
            new_rows_extern.append({
                'Pos:': pos,
                'Bezeichnung': description,
                'Menge': amount,
                'Einzelpreis / EUR netto': None,
                'Gesamtpreis / EUR netto': None
            })

    # convert into an data frame
    df_cost_estimate_extern = pd.DataFrame(new_rows_extern)

    return df_cost_estimate_extern


def excel_export_tender_intern(#excel_path: str,
                               df_prepared: pd.DataFrame) -> None:
    """Export tender data to a structured Excel file with formatted styling (intern).

    Parameters: 
        excel_path: file path where the Excel file will be saved.
        df_prepared : pepared Data of furniture for cost estimate.
    
    Returns: The function saves the formatted Excel file to the specified `excel_path` and does not return a value.
    """

    # create new excel document
    wb = Workbook()
    ws = wb.active

    # add title of the project
    ws.append(['Muster Uni'])
    ws.append([])
    ws.append(['Projekt: Musterstadt, Musterstraße 7'])
    ws.append(['Los 1 - Büromobiliar'])
    ws.append([])

    # set style: border
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    # set style
    header_font = Font(size=10, color="006100")
    default_font = Font(size=10)
    vertical_top_alignment = Alignment(vertical='top', wrap_text=True)

    # write data frame into the excel file (starting from the 6th row)
    for r_idx, r in enumerate(dataframe_to_rows(df_prepared, index=False, header=True), 1):
        ws.append(r)

        # set format
        for cell in ws[r_idx + 5]:
            cell.border = thin_border
            cell.font = default_font
            cell.alignment = vertical_top_alignment

        # style header of table
        if r_idx == 1:
            for cell in ws[r_idx + 5]:
                cell.border = thin_border
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                cell.font = header_font
                cell.alignment = vertical_top_alignment

        # format spesific cells (pos, amount, single_price)
        if r_idx + 5 > 5:
            ws.cell(row=r_idx + 5, column=1).number_format = '@'
            ws.cell(row=r_idx + 5, column=5).number_format = '0'
            ws.cell(row=r_idx + 5, column=6).number_format = '0.00'

            # formular for total_price
            if r_idx > 1:
                ws.cell(row=r_idx + 5, column=7).value = f"=E{r_idx + 5}*F{r_idx + 5}"
                ws.cell(row=r_idx + 5, column=7).number_format = '0.00'

            # enable text wrapping for the description
            description_cell = ws.cell(row=r_idx + 5, column=3)
            description_cell.alignment = vertical_top_alignment
            description_cell.font = default_font

            # color filling for the column "single_price"
            single_price_cell = ws.cell(row=r_idx + 5, column=6)
            if r_idx > 1:
                single_price_cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    # set font of headline
    for row in range(1, 5):
        for cell in ws[row]:
            cell.font = Font(name="Calibri", size=14)

    # setting column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 60
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 30
    ws.column_dimensions['G'].width = 30

    # set zoom level and hide gridlines
    ws.sheet_view.zoomScale = 80
    ws.sheet_view.showGridLines = False

    # save as excel file
    # wb.save(excel_path)
    
    # Sent data to stream
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0) 

    return excel_stream


# Export the final tender as excel file
def excel_export_tender_extern(#excel_path: str,
                               df_prepared: pd.DataFrame) -> None:
    """Export tender data to a structured Excel file with formatted styling (extern).

    Parameters: 
        excel_path: file path where the Excel file will be saved.
        df_prepared : pepared Data of furniture for cost estimate.
    
    Returns: The function saves the formatted Excel file to the specified `excel_path` and does not return a value.
    """

    # create new word document
    wb = Workbook()
    ws = wb.active

    # add title of the project
    ws.append(['Muster Uni'])
    ws.append([])
    ws.append(['Projekt: Musterstadt, Musterstraße 7'])
    ws.append(['Los 1 - Büromobiliar'])
    ws.append([])

    # set style: border
    thin_border = Border(left=Side(style='thin'), 
                        right=Side(style='thin'), 
                        top=Side(style='thin'), 
                        bottom=Side(style='thin'))

    # set style
    header_font = Font(size=10, color="006100")
    default_font = Font(size=10)
    vertical_top_alignment = Alignment(vertical='top', wrap_text=True)

    # write data frame into the excel file (starting from the 6th row)
    for r_idx, r in enumerate(dataframe_to_rows(df_prepared, index=False, header=True), 1):
        ws.append(r)

        # set format
        for cell in ws[r_idx + 5]:
            cell.border = thin_border
            cell.font = default_font
            cell.alignment = vertical_top_alignment

        # style header of table
        if r_idx == 1:
            for cell in ws[r_idx + 5]:
                cell.border = thin_border
                cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                cell.font = header_font
                cell.alignment = vertical_top_alignment

        # format spesific cells (pos, amount, single_price)
        if r_idx + 5 > 5:
            ws.cell(row=r_idx + 5, column=1).number_format = '@'
            ws.cell(row=r_idx + 5, column=3).number_format = '0'
            ws.cell(row=r_idx + 5, column=4).number_format = '0.00'

            # formular for total_price
            if r_idx > 1:
                ws.cell(row=r_idx + 5, column=5).value = f"=C{r_idx + 5}*D{r_idx + 5}"
                ws.cell(row=r_idx + 5, column=5).number_format = '0.00'

            # enable text wrapping for the description
            description_cell = ws.cell(row=r_idx + 5, column=2)
            description_cell.alignment = vertical_top_alignment
            description_cell.font = default_font

            # color filling for the column "single_price"
            single_price_cell = ws.cell(row=r_idx + 5, column=4)
            if r_idx > 1:
                single_price_cell.fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    # set font of headline
    for row in range(1, 5):
        for cell in ws[row]:
            cell.font = Font(name="Calibri", size=14)

    # setting column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 30

    # set zoom level and hide gridlines
    ws.sheet_view.zoomScale = 80
    ws.sheet_view.showGridLines = False

    # save as excel file
    #wb.save(excel_path)

    # Sent data to stream
    excel_stream = BytesIO()
    wb.save(excel_stream)
    excel_stream.seek(0) 

    return excel_stream
